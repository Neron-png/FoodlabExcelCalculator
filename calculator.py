from openpyxl import Workbook, load_workbook
import os
from postProcessing import *

def Excel_calculator(input_file, options):


    offset_error = 0.1



    ####################
    # Finding the file #
    ####################
    if input_file is None:
        return "The input file can't be empty."

    for file in os.listdir():
        if "{}.xlsx".format(input_file) == file:
            break
    else:
        return "Could not find {}".format(input_file)


    ##########################
    # Loading the input file #
    ##########################
    wb = load_workbook(filename="{}.xlsx".format(input_file))
    ws = wb.active
    work_area = ['B', 2, 29]
    max_row = ws.max_row
    max_column = ws.max_column
    ref_area = ('G', 2, max_row)
    EOF = max_row

    ###############################
    # Initializing a new workbook #
    ###############################
    new_wb = Workbook()
    new_ws = new_wb.active
    for i in range(max_column):
        for j in range(max_row):
            new_ws[chr(i+65) + str(j+1)] = ws[chr(i+65) + str(j+1)].value


    #################################
    # Creating the output file name #
    #################################
    exit_file = "{}_out.xlsx".format(input_file)
    while exit_file in os.listdir():
        if exit_file[-6] == "t":
            exit_file = exit_file[:-5:] + "0" + exit_file[-5:]
        else:
            exit_file = exit_file[:-6:] + str( int(exit_file[-6]) + 1 ) + exit_file[-5:]


    #######################################
    # Finding the size for the main table #
    #######################################
    i = work_area[1]
    while ws[work_area[0]+str(i)].value is not None and i <= max_row:
        i += 1

    work_area[2] = i


    ########################################################
    # Looping through the cells and doing the comparissons #
    ########################################################
    for i in range(work_area[1], work_area[2]+1):
        cell_value = ws[work_area[0] + str(i)].value
        matching_id = ""
        if cell_value is None: continue

        for j in range(ref_area[1], ref_area[2]+1):
            ref_col = ref_area[0]
            ref_cell = ws[ref_col + str(j)].value

            if ref_cell is None:
                continue
            elif isinstance(ref_cell, str):
                ref_cell = float(ref_cell.replace(",","."))
            elif isinstance(ref_cell, int):
                ref_cell = float(ref_cell/1000)


            if (ref_cell + offset_error) >= cell_value >= (ref_cell - offset_error):
                matching_id = ws[chr(ord(ref_area[0])-1) + str(j)].value
                break

        new_ws[chr(ord(work_area[0])-1) + str(i)] = matching_id


    ###################
    # Post Processing #
    ###################



    ###################################
    # Calculating the saturated cells #
    ###################################
    if options.calcSaturated.get() == True:
        saturated(ws, new_ws, work_area, max_row)

    if options.calcMono.get() == True:
        mono(ws, new_ws, work_area, max_row)

    if options.calcTrans.get() == True:
        trans(ws, new_ws, work_area, max_row)

    if options.calcPoly.get() == True:
        poly(ws, new_ws, work_area, max_row)




    new_wb.save(exit_file)


    return str("Saved to {}".format(exit_file))

